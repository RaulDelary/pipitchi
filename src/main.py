import sys
from datetime import datetime
from openpyxl import load_workbook
from emoji import emojize
from requests import post
from json import dumps
from configparser import ConfigParser

ms_teams_webhook = None

try:
  file_name = sys.argv [1]
  wb = load_workbook (filename = file_name, data_only = True)
  configs = ConfigParser ()
  configs.read (filenames = './app.ini', encoding = 'UTF-8')

  ms_teams_webhook = configs ['connectors']['teams_webhook']


except FileNotFoundError:
  print (f'Não foi possível abrir o arquivo {file_name}. Verifique se o caminho do arquivo está correto.')
  exit (1)

except KeyError:
  print (f'Não foi possível carregar o arquivo de configuração. Webhook não será utilizado')

except IndexError:
  # To Do: Implement usage function.
  print (f'Nome do arquivo de entrada não foi informado.')
  exit (1)


msg_template = '''
{em01} Health-Check Backup
{em02} {data_hora}

{em03} Infraestrutura
(AVAMAR E VEEAM)

RJ1
  {em04} Área Ocupada: {area_ocupada_rj1}%
  {em05} Área Livre: {area_livre_rj1} TB
  {em06} {cres_dec_str_rj1}: {cres_dec_value_rj1} TB

RJ2
  {em07} Área Ocupada: {area_ocupada_rj2}%
  {em08} Área Livre: {area_livre_rj2} TB
  {em09} {cres_dec_str_rj2}: {cres_dec_value_rj2} TB

{em10} Replicação AWS
  {em11} RJ1 Executado: {rep_aws_executado_rj1}%
  {em12} Dt última: {rep_aws_ultima_rj1}
  {em13} Área Livre: {rep_aws_livre_rj1}%

  {em14} RJ2 executado: {rep_aws_executado_rj2}%
  {em15} Dt última: {rep_aws_ultima_rj2}
  {em16} Área Livre: {rep_aws_livre_rj2}%


{em17} Execuções Data Center
  {em18} JOBS RJ1: {jobs_rj1}
  {em19} Executado: {jobs_executado_rj1:.2f}% 

  {em20} JOBS RJ2: {jobs_rj2}
  {em21} Executado: {jobs_executado_rj2:.2f}%


{em22} Execuções Unidades
  {em23} Escopo: {count_units} Locais
  {em24} Sucesso: {pct_sucesso_total:.2f}%
'''


ws = wb.active
status_column = tuple (ws.iter_cols (min_col = 4, max_col = 4, min_row = 8, values_only = True)) [0]
count_units = len (status_column)
count_valid_units = len ([x for x in status_column if x != 'Sem backup no dia'])
count_success = len ([x for x in status_column if x == 'Sucesso'])
unit_success_pct = (count_success / count_units) * 100


def conditional_emoji (rule, value = 0):
  match rule:
    case 1:
      return ':check_mark_button:' if value <= 79 else ':warning:' if value > 79 and value <= 90 else ':hollow_red_circle:'

    case 2:
      return ':hollow_red_circle:' if value <= 80 else ':warning:' if value > 80 and value <= 90 else ':check_mark_button:'

    case 3:
      return ':hollow_red_circle:' if value <= 10 else ':warning:' if value > 10 and value <= 19 else ':check_mark_button:'

    case 4:
      return ':check_mark_button:' if value >= 95 else ':warning:'

    case 5:
      return ':hollow_red_circle:' if 'Falha' in status_column else ':warning:' if 'Parcial' in status_column else ':check_mark_button:'

def conditional_string (rule, value = 0):
  match rule:
    case 1:
      return 'Crescimento/dia' if value >= 0 else 'Redução nas últimas 24h'
    
    case 2:
      date_str = value.split ('\n') [0].split () [2]
      return f'{date_str [8:]}/{date_str [5:7]}/{date_str [0:4]}'

def send_teams_message (message):
  try:
    payload = {'text': message}
    
    if (post (ms_teams_webhook, data = dumps (payload)).status_code != 200):
      raise Exception ('Status Code da requisição diferente de 200.')

  except:
    print ('Não foi possível utilizar o webhook. Verifique se o endpoint está correto ou suas configurações de rede.')

output_text = msg_template.format (
  em01 = ':stethoscope:',
  em02 = ':spiral_calendar:', data_hora = datetime.now ().strftime ('%d/%m/%Y'),

  em03 = ':large_blue_diamond:',
  em04 = conditional_emoji (1, ws ['O3'].value * 100), area_ocupada_rj1 = int (ws ['O3'].value * 100),
  em05 = ':information:', area_livre_rj1 = ws ['I3'].value,
  em06 = ':information:', cres_dec_str_rj1 = conditional_string (1, ws ['J3'].value), cres_dec_value_rj1 = ws ['J3'].value,

  em07 = conditional_emoji (1, ws ['O4'].value * 100), area_ocupada_rj2 = int (ws ['O4'].value * 100),
  em08 = ':information:', area_livre_rj2 = ws ['I4'].value,
  em09 = ':information:', cres_dec_str_rj2 = conditional_string (1, ws ['J4'].value), cres_dec_value_rj2 = ws ['J4'].value,

  em10 = ':large_blue_diamond:',
  em11 = conditional_emoji (2, ws ['G3'].value * 100), rep_aws_executado_rj1 = ws ['G3'].value * 100,
  em12 = ':information:', rep_aws_ultima_rj1 = conditional_string (2, ws ['H3'].value),
  em13 = conditional_emoji (3, ws ['K3'].value), rep_aws_livre_rj1 = ws ['K3'].value,

  em14 = conditional_emoji (2, ws ['G4'].value * 100), rep_aws_executado_rj2 = ws ['G4'].value * 100,
  em15 = ':information:', rep_aws_ultima_rj2 = conditional_string (2, ws ['H4'].value),
  em16 = conditional_emoji (3, ws ['K4'].value), rep_aws_livre_rj2 = ws ['K4'].value,

  em17 = ':large_blue_diamond:',
  em18 = ':information:', jobs_rj1 = ws ['N3'].value,
  em19 = conditional_emoji (4, ws ['E3'].value * 100), jobs_executado_rj1 = ws ['E3'].value * 100,

  em20 = ':information:', jobs_rj2 = ws ['N4'].value,
  em21 = conditional_emoji (4, ws ['E4'].value * 100), jobs_executado_rj2 = ws ['E4'].value * 100,

  em22 = ':large_blue_diamond:',
  em23 = conditional_emoji (5), count_units = count_units,
  em24 = ':information:', pct_sucesso_total = unit_success_pct
)




emojized_output = emojize (output_text)
payload = {'text': emojized_output}


print (emojized_output)

if ms_teams_webhook:
  send_teams_message (emojized_output)